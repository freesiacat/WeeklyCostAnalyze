#============================================#
#
#  担当別で工数を週単位で集計する
#  作成者    ：セイカ
#  作成日    ：2024/10/27
#  バージョン：v1.00
#
#============================================#

#★import群==================================#
import openpyxl      #excel操作用
from openpyxl.styles.borders import Border, Side
import calendar      #日付操作用
import pandas as pd  #データ分析用
import tkinter as tk #GUI用
import tkinter.filedialog
from tkinter import messagebox
import os            #ファイル存在チェック用
import sys


#★定数群====================================#
#チェック用シート名
SHEET_CHECK = "チェック用"
#集計結果用シート名
SHEET_SUM   = "集計結果"
#テンプレートファイル名
TEMPLATE_NAME = "工数表.xlsx"
#アプリバージョン
APP_VERSION = "v1.00"
#担当項目
COLUMN_TANTO_NM     = "担当"
#開始日項目
COLUMN_START_DT_CHECK = "A" #「チェック用」シート用
COLUMN_START_DT_SUM   = "B" #「集計結果」シート用
COLUMN_START_DT_NM    = "開始日"
#工数項目
COLUMN_CONTS_CHECK    = "B" #「チェック用」シート用
COLUMN_CONTS_SUM      = "C" #「集計結果」シート用
COLUMN_CONTS_NM     = "工数"
#Excel最大行数
EXCEL_ROWS_MAX = 1048576
#可用連続空行数(設定行以上スキップがある場合はチェックやめる)
EXCEL_NONE_ROWS = 10

#★関数群====================================#
# 集計処理実行
def pushInputPash():
    strErrMessage = None #エラー時のメッセージ用
    inputPath=txtInputDir.get()

    #●エラーチェック～～～～～～～～～～～～～～～
    #ファイル存在チェック
    if not os.path.isfile(inputPath):
        strErrMessage = "パスが存在しません！ \n" + inputPath
        messagebox.showerror("!!ファイル存在チェック!!", strErrMessage)
        return None
    
    #拡張子のチェック
    if not inputPath.endswith(".xlsx"):
        strErrMessage = "有効な拡張子(.xlsx)ではありません！ \n" + inputPath
        messagebox.showerror("!!ファイル拡張子チェック!!", strErrMessage)
        return None

    #ファイルが開かれているかどうかチェック
    if xlsx_is_open(inputPath):
        strErrMessage = "該当のファイルが開かれています！ \n" + inputPath
        messagebox.showerror("!!ファイル編集チェック!!", strErrMessage)
        return None
    
    #「チェック用」シートの存在チェック
    #既存の「集計結果」シートを削除する
    wb = openpyxl.load_workbook(inputPath)
    isCheckSheet = False
    # シートをループ
    for ws in wb.worksheets:
        #「チェック用」シートチェック
        if ws.title == SHEET_CHECK:
            isCheckSheet = True
        # 既存の「集計結果」シートを削除
        if ws.title == SHEET_SUM:
            wb.remove(ws)
    wb.save(inputPath)
    
    #「チェック用」シートが存在しない場合はエラー
    if not isCheckSheet:
        strErrMessage = "チェック対象の「" + SHEET_CHECK + "」シートが存在しません！" +  "\n" + inputPath
        messagebox.showerror("!!シート存在チェック!!", strErrMessage)
        return None

    #シート設定(チェック用)
    ws = wb[SHEET_CHECK]

    #チェック行カウント用
    intCountRow = 1
    #連続空行数カウント用
    intCountNoneRows = 0 
    #空行の行数
    intNoneRows = 1
    #初期化
    tmpColumn_stdt = None
    tmpColumn_cost = None
    while EXCEL_ROWS_MAX>intCountRow:
        tmpColumn_stdt = COLUMN_START_DT_CHECK  + str(intCountRow)
        tmpColumn_cost = COLUMN_CONTS_CHECK     + str(intCountRow)
        #「開始日」に値が入っていないか確認
        if isCellBlank(ws, tmpColumn_stdt):
            intNoneRows = intCountRow
            intCountNoneRows = intCountNoneRows + 1
            #「工数」に値が入っている場合はエラー
            if not isCellBlank(ws, tmpColumn_cost):
                strErrMessage = "★「" + SHEET_CHECK + "」シート \n"
                strErrMessage = strErrMessage + "「" + COLUMN_START_DT_NM + "」が入力されていない行" 
                strErrMessage = strErrMessage + "("  + tmpColumn_stdt + ")があります。" + "\n"
                strErrMessage = strErrMessage + "日付を入力するか行を削除してください！"
                messagebox.showerror("!!項目チェック!!", strErrMessage)
                return None
        else:
            #途中に空行がある場合はエラーにする
            if intCountNoneRows>0:
                strErrMessage = "★「" + SHEET_CHECK + "」シート \n"
                strErrMessage = strErrMessage + "空行があります！(" + str(intNoneRows) + "行目)" + "\n"
                strErrMessage = strErrMessage + "空行を削除してください！"
                messagebox.showerror("!!項目チェック!!", strErrMessage)
                return None
            #タイトル行以降をチェック
            if intCountRow > 1:
                #開始日列に日付以外が入っている場合はエラーにする
                if not ws[tmpColumn_stdt].data_type =="d":
                    strErrMessage = "★「" + SHEET_CHECK + "」シート \n"
                    strErrMessage = strErrMessage + "「" + COLUMN_START_DT_NM + "」に"
                    strErrMessage = strErrMessage + "日付以外が入力されています！(" + tmpColumn_cost + ")" + "\n" + "日付を入力してください！"
                    messagebox.showerror("!!項目チェック!!", strErrMessage)
                    return None
                #工数列に数値以外が入っている場合はエラーにする
                if not isCellBlank(ws, tmpColumn_cost) and not ws[tmpColumn_cost].data_type =="n":
                    strErrMessage = "★「" + SHEET_CHECK + "」シート \n"
                    strErrMessage = strErrMessage + "「" + COLUMN_CONTS_NM + "」に"
                    strErrMessage = strErrMessage + "数値以外が入力されています！(" + tmpColumn_cost + ")" + "\n" + "数値を入力してください！"
                    messagebox.showerror("!!項目チェック!!", strErrMessage)
                    return None

        #設定行連続で空行になった場合は、ループから抜ける(全行チェックすると時間かかるので・・・)
        if intCountNoneRows>EXCEL_NONE_ROWS:
            break
        intCountRow = intCountRow + 1

    #●データ分析処理～～～～～～～～～～～～～～～
    dfWbs = pd.read_excel(inputPath, sheet_name=SHEET_CHECK)
    dfWbsSum=dfWbs.groupby(COLUMN_TANTO_NM).resample("W",on=COLUMN_START_DT_NM, label="left",closed="left").sum()
    #別シートへ出力処理
    with pd.ExcelWriter(inputPath,mode='a') as writer:
        dfWbsSum.iloc[:,0:1].to_excel(writer, sheet_name=SHEET_SUM)
    
    #出力結果の書式変更
    wb = openpyxl.load_workbook(inputPath)
    #シート設定(集計結果)
    ws = wb[SHEET_SUM]
    # 列全体に書式を適用
    # 開始日列は「yyyy/mm/dd」形式
    for cell in ws[COLUMN_START_DT_SUM]:
       cell.number_format = "yyyy/mm/dd"
    # 工数列は「0.00」形式
    for cell in ws[COLUMN_CONTS_SUM]:
       cell.number_format = "0.00"
    #列幅変更
    ws.column_dimensions[COLUMN_START_DT_SUM].width = 13
    # 値が存在するセルに順番に罫線を引く(上下左右引く)
    side = Side(style="thin", color="000000")
    border = Border(top=side, bottom=side, right=side, left=side)
    for row in ws:
        for cell in row:
            # cell.coordinateでセルの番地を"A3"とセルアドレスに変換
            ws[cell.coordinate].border = border
    wb.save(inputPath)
    
    messagebox.showinfo("!!処理結果!!", "集計処理完了しました！　\n " + "「" + SHEET_SUM + "」シートを確認してください。")


# ファイル選択画面
def searchFilePath():
    fTyp = [("Excelファイル","xlsx"),("","*")]
    iDir = os.path.abspath(os.path.dirname(__file__))
    file = tk.filedialog.askopenfilename(filetypes = fTyp,initialdir = iDir)
    txtInputDir.delete(0, tk.END)
    txtInputDir.insert(tkinter.END, file)

#ファイルを追記モードで開けるかチェック
def xlsx_is_open(filepath: str) -> bool:
    try:
        f = open(filepath, 'a')
        f.close()
    except:
        return True
    else:
        return False
    
#値の存在チェック
def isCellBlank(ws, cellCoodinate):
    if ws[cellCoodinate].value is None:
        return True
    else:
        return False


#★初期処理==================================#
# tkオブジェクトの作成
root = tk.Tk()
root.title("★週別工数分析ツール " + APP_VERSION) #ウィンドウのタイトルを設定
root.geometry("600x80")     #ウィンドウのサイズを設定
root.resizable(False, False) #サイズ変更不可
#　実行時の初期ディレクトリ取得
#base_path = os.path.dirname(__file__) + "\\" + TEMPLATE_NAME
base_path = os.path.dirname(sys.argv[0]) + "\\" + TEMPLATE_NAME



# 画面部品定義================================#
# チェック用パス登録用---------------*
lblInputDir = tk.Label(root, text="チェックExcelパス：")
lblInputDir.place(x=10, y=10)
txtInputDir = tk.Entry(width=70)
txtInputDir.place(x=110, y=10)
txtInputDir.delete(0, tk.END)
txtInputDir.insert(tkinter.END, base_path)
btnSearchPath = tk.Button(root, text="選択", command=searchFilePath)
btnSearchPath.place(x=550, y=5)
btnInputPath = tk.Button(root, text="集計する", command=pushInputPash)
btnInputPath.place(x=485, y=40, width=100)

# メインループの実行
root.mainloop()
